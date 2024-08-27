import pandas as pd
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from openpyxl import load_workbook
import os

# Inicializa DataFrames vacíos
libro_diario = pd.DataFrame(columns=['Fecha', 'Cuenta', 'Débito', 'Crédito', 'Descripción'])
balance_mayor = pd.DataFrame(columns=['Cuenta', 'Débito', 'Crédito', 'Saldo'])

def agregar_transaccion(fecha, cuenta, debito, credito, descripcion):
    global libro_diario
    nueva_transaccion = pd.DataFrame([[fecha, cuenta, debito, credito, descripcion]], columns=libro_diario.columns)
    libro_diario = pd.concat([libro_diario, nueva_transaccion], ignore_index=True)
    actualizar_balance_mayor()
    guardar_en_excel(fecha, cuenta, debito, credito, descripcion)

def actualizar_balance_mayor():
    global balance_mayor
    balance_mayor = libro_diario.groupby('Cuenta').agg({
        'Débito': 'sum',
        'Crédito': 'sum'
    }).reset_index()
    balance_mayor['Saldo'] = balance_mayor['Débito'] - balance_mayor['Crédito']

def ver_libro_diario():
    global libro_diario
    return libro_diario.to_string(index=False)

def ver_balance_mayor():
    global balance_mayor
    return balance_mayor.to_string(index=False)

def guardar_en_excel(fecha, cuenta, debito, credito, descripcion):
    file_path = 'libro_diario.xlsx'
    
    if os.path.exists(file_path):
        # Si el archivo existe, lo carga y agrega los nuevos datos
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book
    else:
        # Si no existe, lo crea
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
    
    # Cargar la hoja actual o crearla si no existe
    if 'Libro Diario' in writer.book.sheetnames:
        libro_diario_df = pd.read_excel(file_path, sheet_name='Libro Diario')
    else:
        libro_diario_df = pd.DataFrame(columns=['Fecha', 'Cuenta', 'Débito', 'Crédito', 'Descripción'])
    
    # Agregar la nueva transacción al DataFrame existente
    nueva_transaccion = pd.DataFrame([[fecha, cuenta, debito, credito, descripcion]], columns=libro_diario_df.columns)
    libro_diario_df = pd.concat([libro_diario_df, nueva_transaccion], ignore_index=True)
    
    # Guardar el DataFrame actualizado en la hoja de Excel
    libro_diario_df.to_excel(writer, sheet_name='Libro Diario', index=False)
    writer.save()

class ContableApp(App):
    def build(self):
        layout = BoxLayout(orientation='vertical')
        
        # Entrada de datos
        input_layout = GridLayout(cols=2)
        
        # Etiquetas para los campos de entrada
        input_layout.add_widget(Label(text='Fecha:'))
        self.fecha_input = TextInput(hint_text='Fecha', multiline=False)
        input_layout.add_widget(self.fecha_input)
        
        input_layout.add_widget(Label(text='Cuenta:'))
        self.cuenta_input = TextInput(hint_text='Cuenta', multiline=False)
        input_layout.add_widget(self.cuenta_input)
        
        input_layout.add_widget(Label(text='Débito:'))
        self.debito_input = TextInput(hint_text='Débito', multiline=False)
        input_layout.add_widget(self.debito_input)
        
        input_layout.add_widget(Label(text='Crédito:'))
        self.credito_input = TextInput(hint_text='Crédito', multiline=False)
        input_layout.add_widget(self.credito_input)
        
        input_layout.add_widget(Label(text='Descripción:'))
        self.descripcion_input = TextInput(hint_text='Descripción', multiline=False)
        input_layout.add_widget(self.descripcion_input)
        
        add_button = Button(text='Agregar Transacción')
        add_button.bind(on_press=self.agregar_transaccion)
        
        layout.add_widget(input_layout)
        layout.add_widget(add_button)
        
        self.result_label = Label(text='No hay datos', size_hint_y=None, height=200)
        layout.add_widget(self.result_label)
        
        return layout

    def agregar_transaccion(self, instance):
        try:
            fecha = self.fecha_input.text
            cuenta = self.cuenta_input.text
            debito = float(self.debito_input.text or 0)
            credito = float(self.credito_input.text or 0)
            descripcion = self.descripcion_input.text
            
            # Validación simple de entrada
            if not fecha or not cuenta or (debito == 0 and credito == 0):
                self.result_label.text = "Por favor, complete todos los campos obligatorios."
                return
            
            agregar_transaccion(fecha, cuenta, debito, credito, descripcion)
            self.result_label.text = f"Libro Diario:\n{ver_libro_diario()}\n\nBalance Mayor:\n{ver_balance_mayor()}"
        except ValueError:
            self.result_label.text = "Error en los valores ingresados. Asegúrese de que Débito y Crédito sean números."

if __name__ == '__main__':
    ContableApp().run()
